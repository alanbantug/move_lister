import tkinter
from tkinter import *

from tkinter.ttk import *
from tkinter import messagebox

from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
from tkcalendar import Calendar
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment

from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles import PatternFill, Font, Alignment

import os
import json
import psycopg2


class Application(Frame):

    def __init__(self, master):

        self.master = master
        self.main_container = Frame(self.master)

        # Define the file locations
        self.target = StringVar()
        self.tag = StringVar()
        self.sheet = StringVar()
        self.sheet_saved = StringVar()
        self.opening = StringVar()
        self.variation = StringVar()
        self.comments = StringVar()
        self.result = IntVar()
        self.whitePlayer = StringVar()
        self.blackPlayer = StringVar()
        self.whiteMove = StringVar()
        self.blackMove = StringVar()
        self.whiteComment = StringVar()
        self.blackComment = StringVar()
        self.type = IntVar()
        self.whitewin = IntVar()
        self.blackwin = IntVar()
        self.sheetsList = ['No Selection ']
        self.allMoves = []
        self.allComments = {}   # dictionary with move number and text
        self.pointer = 0
        self.mode = IntVar()
        self.firstBack = True

        # self.origin = os.getcwd()
        # self.copied = IntVar()
        # self.copying = 0
        # self.source = ""
        # self.target = ""
        # self.script = ""
        # self.allSet = True
        # self.numA = StringVar()
        # self.numB = StringVar()
        # self.numC = StringVar()
        # self.numD = StringVar()
        # self.numE = StringVar()
        # self.numF = StringVar()
        # self.updatedFiles = IntVar()
        # self.initFolders = IntVar()
        

        # Create main frame
        self.main_container.grid(column=0, row=0, sticky=(N,S,E,W))

        # Set Label styles
        
        Style().configure("M.TLabel", font="Courier 20 bold", height="20", foreground="blue", anchor="center")
        Style().configure("A.TLabel", font="Verdana 8")
        Style().configure("D.TLabel", font="Verdana 8", background="white", width=25)
        Style().configure("MS.TLabel", font="Verdana 10" )
        Style().configure("S.TLabel", font="Verdana 8" )
        Style().configure("G.TLabel", font="Verdana 8")
        Style().configure("T.TLabel", font="Verdana 12 bold")
        # Style().configure("B.TLabel", font="Verdana 8", background="white")
        Style().configure("B.TLabel", font="Verdana 8", background="white", width="48")
        Style().configure("C.TLabel", font="Verdana 8", width="12")
        Style().configure("M.TEntry", font="Courier 30")

        # Set button styles
        
        Style().configure("D.TButton", font="Verdana 8", relief="ridge", width=25)
        Style().configure("B.TButton", font="Verdana 8", relief="ridge", width=15)
        Style().configure("E.TButton", font="Verdana 16", relief="ridge")
        Style().configure("B.TRadiobutton", font="Verdana 8", width=7)

        # Set check button styles
        Style().configure("B.TCheckbutton", font="Verdana 8")
        Style().configure("D.TCheckButton", font="Verdana 8", width='20')

        Style().configure("O.TLabelframe.Label", font="Verdana 8", foreground="black")

        self.parentTab = Notebook(self.main_container)
        self.detailTab = Frame(self.parentTab)   # first page, which would get widgets gridded into it
        self.movesTab = Frame(self.parentTab)   # third page
        self.parentTab.add(self.detailTab, text='    Details   ')
        self.parentTab.add(self.movesTab, text='    Moves     ')

        ''' Main container
        '''

        self.mainLabel = Label(self.main_container, text="GAME MOVE ENTRY", style="M.TLabel" )
        self.exit = Button(self.main_container, text="EXIT", style="E.TButton", command=root.destroy)

        self.sep_a = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_b = Separator(self.main_container, orient=HORIZONTAL)

        ''' game selection
        '''
        # self.typeGroup = LabelFrame(self.detailTab, text=' TYPE ', style="O.TLabelframe")
        # self.gameType = Radiobutton(self.typeGroup, text="Game", style="B.TRadiobutton", variable=self.type, value=1)
        # self.openType = Radiobutton(self.typeGroup, text="Opening", style="B.TRadiobutton", variable=self.type, value=2)

        ''' details widgets
        '''
        self.det_a = Separator(self.detailTab, orient=HORIZONTAL)
        self.det_b = Separator(self.detailTab, orient=HORIZONTAL)
        self.det_c = Separator(self.detailTab, orient=HORIZONTAL)
        self.det_d = Separator(self.detailTab, orient=HORIZONTAL)
        self.det_e = Separator(self.detailTab, orient=HORIZONTAL)
        self.det_f = Separator(self.detailTab, orient=HORIZONTAL)

        self.targetOption = LabelFrame(self.detailTab, text=' FILE ', style="O.TLabelframe")
        self.selectTarget = Button(self.targetOption, text="SELECT", style="B.TButton", command=self.setTarget)
        self.targetLabel = Label(self.targetOption, text="", style="B.TLabel" )
        self.sheetOptions = LabelFrame(self.detailTab, text=' SHEET ', style="O.TLabelframe")
        # self.selectSheet = OptionMenu(self.sheetOptions, self.sheet, *self.sheetsList, command=self.getSheetIdsList)
        self.selectSheet = OptionMenu(self.sheetOptions, self.sheet, *self.sheetsList)
        self.tagOptions = LabelFrame(self.detailTab, text=' IDENTIFIER ', style="O.TLabelframe")
        self.tagEntry = Entry(self.tagOptions, textvariable=self.tag, font="Courier 10", width="10")
        

        ''' game information
        '''

        self.gameDetails = LabelFrame(self.detailTab, text=' GAME DETAIL', style="O.TLabelframe")
        self.openLabel = Label(self.gameDetails, text="Opening", style="C.TLabel" )
        self.gameOpening = Entry(self.gameDetails, textvariable=self.opening, font="Courier 10", width="50")
        self.varLabel = Label(self.gameDetails, text="Variation", style="C.TLabel" )
        self.gameVariation = Entry(self.gameDetails, textvariable=self.variation, font="Courier 10", width="50")
        self.whitePlayerL = Label(self.gameDetails, text="White PLayer", style="C.TLabel" )
        self.whitePlayerE = Entry(self.gameDetails, textvariable=self.whitePlayer, font="Courier 10", width="50")
        self.blackPlayerL = Label(self.gameDetails, text="Black Player", style="C.TLabel" )
        self.blackPlayerE = Entry(self.gameDetails, textvariable=self.blackPlayer, font="Courier 10", width="50")

        self.getDetails = Button(self.detailTab, text="GET DETAILS", style="B.TButton", command=self.getGame)
        self.saveDetails = Button(self.detailTab, text="SAVE ", style="B.TButton", command=self.saveGameDetails)
        self.resetEntry = Button(self.detailTab, text="RESET", style="B.TButton", command=self.resetEntryFields)

        ''' game moves
        '''
        self.mov_a = Separator(self.movesTab, orient=HORIZONTAL)
        self.mov_b = Separator(self.movesTab, orient=HORIZONTAL)
        self.mov_c = Separator(self.movesTab, orient=HORIZONTAL)
        self.mov_d = Separator(self.movesTab, orient=HORIZONTAL)

        self.comFrame = LabelFrame(self.movesTab, text=' Comments', style="O.TLabelframe")
        self.gameComments = Text(self.comFrame, width="44", height="3")
        self.resFrame = LabelFrame(self.movesTab, text=' Results', style="O.TLabelframe")
        self.whiteWinner = Checkbutton(self.resFrame, text='White wins ', style="B.TCheckbutton", variable=self.whitewin)
        self.blackWinner = Checkbutton(self.resFrame, text='Black wins ', style="B.TCheckbutton", variable=self.blackwin)
        
        self.gameMoves = LabelFrame(self.movesTab, text=' Game Moves', style="O.TLabelframe")
        self.whiteMoveFrame = LabelFrame(self.gameMoves, text=' White', style="O.TLabelframe")
        self.blackMoveFrame = LabelFrame(self.gameMoves, text=' Black', style="O.TLabelframe")
        self.whiteMoveE = Entry(self.whiteMoveFrame, textvariable=self.whiteMove, font="Courier 34 bold", width="9")
        self.blackMoveE = Entry(self.blackMoveFrame, textvariable=self.blackMove, font="Courier 34 bold", width="9")
        self.whiteComL = Label(self.whiteMoveFrame, text="Comments", style="C.TLabel" )
        self.blackComL = Label(self.blackMoveFrame, text="Comments", style="C.TLabel" )
        self.whiteCom = Text(self.whiteMoveFrame, width="30", height="3")
        self.blackCom = Text(self.blackMoveFrame, width="30", height="3")

        self.next = Button(self.gameMoves, text="NEXT", style="B.TButton", command=self.getNextPair)
        self.prev = Button(self.gameMoves, text="PREV", style="B.TButton", command=self.getPrevPair)

        ''' position widgets
        '''
        self.mainLabel.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        self.parentTab.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky='NSEW')

        ''' details
        '''

        self.selectTarget.grid(row=0, column=0, columnspan=2, padx=5, pady=(5,10), sticky='NSEW')
        self.targetLabel.grid(row=0, column=2, columnspan=2, padx=10, pady=(5,10), sticky='NSEW')
        self.targetOption.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.selectSheet.grid(row=0, column=0, padx=(5,10), pady=(5,10), sticky='NSEW')
        self.selectSheet.config(width=35)
        self.sheetOptions.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.tagEntry.grid(row=0, column=0, padx=10, pady=(5,10), sticky='NSEW')
        self.tagOptions.grid(row=5, column=2, columnspan=1, padx=5, pady=5, sticky='NSEW')
        
        self.resetEntry.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        self.det_a.grid(row=7, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.openLabel.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.gameOpening.grid(row=0, column=2, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.varLabel.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.gameVariation.grid(row=1, column=2, columnspan=2, padx=5, pady=5, sticky='NSEW')
        
        self.whitePlayerL.grid(row=2, column=0, columnspan=1, padx=5, pady=5, sticky='NSEW')
        self.whitePlayerE.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky='NSEW')
        self.blackPlayerL.grid(row=3, column=0, columnspan=1, padx=5, pady=5, sticky='NSEW')
        self.blackPlayerE.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky='NSEW')

        self.gameDetails.grid(row=8, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.det_b.grid(row=9, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.saveDetails.grid(row=10, column=0, padx=5, pady=5, sticky='NSEW')
        self.saveDetails.config(width=38)
        self.getDetails.grid(row=10, column=2, padx=5, pady=5, sticky='NSEW')
        self.getDetails.config(width=38)
        

        ''' moves
        '''

        self.gameComments.grid(row=0, column=0, rowspan=3, padx=5, pady=(5,10), sticky='NSEW')
        self.comFrame.grid(row=0, column=0, columnspan=3, padx=5, pady=5, sticky='NSEW')

        self.whiteWinner.grid(row=0, column=0, padx=5, pady=5, sticky='NSEW')
        self.blackWinner.grid(row=1, column=0, padx=5, pady=5, sticky='NSEW')
        self.resFrame.grid(row=0, column=3, columnspan=1, padx=5, pady=5, sticky='NSEW')

        self.mov_a.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.whiteMoveE.grid(row=0, column=0, columnspan=4, padx=(5,10), pady=(5,10), sticky='NSEW')
        self.whiteComL.grid(row=1, column=0, columnspan=4, padx=(5,10), pady=5, sticky='NSEW')
        self.whiteCom.grid(row=2, column=0, columnspan=4, padx=(5,10), pady=(5,10), sticky='NSEW')
        self.whiteMoveFrame.grid(row=0, column=0, padx=5, pady=5, sticky='NSEW')
        self.blackMoveE.grid(row=0, column=0, columnspan=4, padx=(5,10), pady=(5,10), sticky='NSEW')
        self.blackComL.grid(row=1, column=0, columnspan=4, padx=(5,10), pady=5, sticky='NSEW')
        self.blackCom.grid(row=2, column=0, columnspan=4, padx=(5,10), pady=(5,10), sticky='NSEW')
        self.blackMoveFrame.grid(row=0, column=2, padx=5, pady=5, sticky='NSEW')
        self.prev.grid(row=1, column=0, padx=5, pady=5, sticky='NSEW')
        self.next.grid(row=1, column=2, padx=5, pady=5, sticky='NSEW')
        self.gameMoves.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.mov_b.grid(row=3, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.sep_a.grid(row=3, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.exit.grid(row=4, column=0, columnspan=4, padx=5, pady=(5,10), sticky='NSEW')

        self.mode.set(0)

    def setTarget(self):

        pathname = askopenfilename()

        try:
            self.target.set(pathname)

            if self.target.get().endswith(".xlsx") or self.target.get().endswith(".xls"):
                self.target.set(pathname)
                self.targetLabel["text"] = self.target.get().split("/")[-1]
                self.getSheetList()

            else:
                messagebox.showerror("Invalid file selected", "Invalid file type was selected. Please select again.")
                self.target.set('')
        except Exception as e:
            pass

    def getSheetList(self):

        wb = load_workbook(self.target.get())

        self.sheetsList = ['No Selection ']

        for ws in wb.worksheets:
            self.sheetsList.append(ws.title)

        menu = self.selectSheet['menu']
        menu.delete(0, 'end')

        for sh in self.sheetsList:
            menu.add_command(label=sh, command=lambda value=sh: self.setSelectSheet(value))

    def setSelectSheet(self, sheet):
        
        self.sheet.set(sheet)
        self.sheet_saved.set(sheet)
 
    def getGame(self):

        if self.getGameDetails():

            self.getGameMoves()
            self.postFirstMove()
            
            self.mode.set(1)

            messagebox.showinfo("Tag found", "Tag found. Details were provided.")
        else:
            messagebox.showerror("Not found", "Tag not found. Please try again.")

    def getGameDetails(self):

        wb = load_workbook(self.target.get())
        ws = wb['Index']

        found = False

        count = 1

        while True:

            cell = 'A' + str(count)

            if ws[cell].value:
                if ws[cell].value == self.tag.get():
                    found = True
                    break

            else:
                break

            count += 1

        if found:

            cell = 'B' + str(count)
            self.opening.set(ws[cell].value)

            cell = 'C' + str(count)
            self.variation.set(ws[cell].value)

            cell = 'D' + str(count)
            self.whitePlayer.set(ws[cell].value)

            cell = 'E' + str(count)
            self.blackPlayer.set(ws[cell].value)

            cell = 'F' + str(count)
            if ws[cell].value == 'B':
                self.blackwin.set(1)
                self.whitewin.set(0)
            elif ws[cell].value == 'W':
                self.blackwin.set(0)
                self.whitewin.set(1)
            else:
                self.blackwin.set(1)
                self.whitewin.set(1)

            return True
        else:
            return False
        
    def getGameMoves(self):

        w_sheet, w_col, b_col = self.locateMoves(1) 

        cell = w_col + '1'
        fgColor = w_sheet[cell].fill.fgColor.index            

        # if fgColor == 'FF00FF00':
        #     self.blackwin.set(0)
        # else:
        #     self.blackwin.set(1)

        self.allMoves = []
        count = 2
        
        while True:
            
            w_cell = w_col + str(count)
            b_cell = b_col + str(count)

            if w_sheet[w_cell].value:
                w_move = w_sheet[w_cell].value 
            else:
                break
                
            if w_sheet[b_cell].value:
                b_move = w_sheet[b_cell].value
            else:
                break

            self.allMoves.append([w_move, b_move])
            count += 1

        self.pointer = 0

        d_cell = b_col + '1'
        
        self.gameComments.delete(1.0, END)

        try:
            self.gameComments.insert(END, w_sheet[d_cell].comment.text.rstrip())
        except:
            self.gameComments.insert(END, 'No description')

    def locateMoves(self, type):

        wb = load_workbook(self.target.get())
        ws = wb[self.sheet.get()]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
        colb = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

        if type == 1:
            for ca, cb in zip(cola, colb):
                cell = ca + '1'

                if ws[cell].value == self.tag.get():    
                    return ws, ca, cb 
        else:
            pass 

    def postFirstMove(self):

        self.pointer = 0

        w_move, b_move = self.allMoves[self.pointer]
        self.whiteMove.set(w_move)
        self.blackMove.set(b_move)

    def getNextPair(self):

        if self.pointer + 1 == len(self.allMoves):
            messagebox.showinfo("Last moves", "Last moves already displayed.")
            return

        if self.mode.get() == 1:
            pass
        else:
            self.storePair()
            return

        self.pointer += 1

        w_move, b_move = self.allMoves[self.pointer]

        self.whiteMove.set(w_move)
        self.blackMove.set(b_move)

    def getPrevPair(self):

        if self.pointer == 0:
            messagebox.showinfo("First moves", "First moves already displayed.")
            return

        self.pointer -= 1

        w_move, b_move = self.allMoves[self.pointer]

        self.whiteMove.set(w_move)
        self.blackMove.set(b_move)

        if self.pointer in self.allComments:
            w_com, b_com = self.allComments[self.pointer]

            self.whiteCom.delete(1.0, END)
            self.whiteCom.insert(END, w_com)

            self.blackCom.delete(1.0, END)
            self.blackCom.insert(END, b_com)

    def storePair(self):

        self.pointer += 1

        w_move = self.whiteMove.get()
        b_move = self.blackMove.get()

        self.allMoves.append([w_move, b_move])
        w_com = self.whiteCom.get("1.0", END).strip()
        b_com = self.blackCom.get("1.0", END).strip()

        if w_com or b_com:
            print('storing')
            self.allComments[self.pointer + 1] = [w_com, b_com]

        self.blackMove.set("")
        self.whiteMove.set("")

    def saveGameDetails(self):

        if self.pointer <= 10:
            messagebox.showerror("Incomplete", "Moves not entered or incomplete")
            return 
        
        if self.checkSlots():

            self.checkEntries()
            self.saveToIndex()
            self.saveGameMoves()

            messagebox.showinfo("Entry saved", "Game details and moves are saved.")

        else:
            messagebox.showerror("Not available", "No available slots for game in sheet.")
            return
    
        self.mode.set(2)

    def checkSlots(self):

        wb = load_workbook(self.target.get())
        ws = wb[self.sheet.get()]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
        colb = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

        ''' first check if tag is foound in sheet
        '''

        for a, b in zip(cola, colb):
            cell = a + '1'

            if ws[cell].value == self.tag.get():
                self.column_w = a
                self.column_b = b
                return True

        ''' check for next empty slot
        '''
        
        for a, b in zip(cola, colb):
            cell = a + '1'

            if ws[cell].value:
                pass 
            else:
                self.column_w = a
                self.column_b = b
                return True

        return False 

    def checkEntries(self):

        if self.whiteMove.get():
            self.storePair()

    def saveToIndex(self):

        wb = load_workbook(self.target.get())
        ws = wb['Index']

        count = 2

        while True:
            
            cell = 'A' + str(count)

            if ws[cell].value:
                pass 
            else:
                break 

            count += 1

        self.alignText(ws[cell], 1)
        ws[cell].value = self.tag.get()

        cell = 'B' + str(count)
        self.alignText(ws[cell], 0)
        ws[cell].value = self.opening.get()
        
        cell = 'C' + str(count)
        self.alignText(ws[cell], 1)
        ws[cell].value = self.variation.get()

        cell = 'D' + str(count)
        self.alignText(ws[cell], 1)
        ws[cell].value = self.whitePlayer.get()

        cell = 'E' + str(count)
        self.alignText(ws[cell], 1)
        ws[cell].value = self.blackPlayer.get()

        cell = 'F' + str(count)
        self.alignText(ws[cell], 1)

        if self.blackwin.get() and self.whitewin.get():
            ws[cell].value = 'D'
        if self.blackwin.get():
            ws[cell].value = 'B'
        else:
            ws[cell].value = 'W'

        wb.save(self.target.get())

    def saveGameMoves(self):

        wb = load_workbook(self.target.get())
        ws = wb[self.sheet.get()]

        count = 1

        cell = self.column_w + str(count)
        self.alignText(ws[cell], 1)
        ws[cell].value = self.tag.get()

        if self.blackwin.get() == 1:
            ws[cell].fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        else:
            ws[cell].fill = PatternFill(start_color="00FF00FF", end_color="00FF00FF", fill_type="solid")

        cell = self.column_b + str(count)
        if self.gameComments.get("1.0", END):
            ws[cell].comment = Comment(self.gameComments.get("1.0", END), "")
        
        count += 1
        for i in range(self.pointer):
            w_move, b_move = self.allMoves[i]

            cell = self.column_w + str(count)
            self.alignText(ws[cell], 1)
            ws[cell].value = w_move 

            cell = self.column_b + str(count)
            self.alignText(ws[cell], 1)
            ws[cell].value = b_move 
            
            count += 1
                        
        for key, value in self.allComments.items():

            w_com, b_com = value 

            if w_com:
                cell = self.column_w + str(key)
                ws[cell].comment = Comment(w_com, "")

            if b_com:
                cell = self.column_b + str(key)
                ws[cell].comment = Comment(b_com, "")


        wb.save(self.target.get())

    def alignText(self, ws_cell, mode):

        ws_cell.font = Font(name="Courier New", sz=10)

        if mode:
            ws_cell.alignment = Alignment(horizontal='center')
        else:
            ws_cell.alignment = Alignment(horizontal='left')

    def resetEntryFields(self):

        self.target.set("")
        self.targetLabel["text"] = ""

        self.sheetsList = ['No Selection ']

        menu = self.selectSheet['menu']
        menu.delete(0, 'end')

        for sh in self.sheetsList:
            menu.add_command(label=sh, command=lambda value=sh: self.setSelectSheet(value))

        self.sheet.set('No Selection')
        
        self.tag.set("")
        self.opening.set("")
        self.variation.set("")
        self.whitePlayer.set("")
        self.blackPlayer.set("")

        self.gameComments.delete(1.0, END)
        self.whiteCom.delete(1.0, END)
        self.blackCom.delete(1.0, END)

        self.pointer = 0
        self.allMoves = []
        self.allComments = {}


root = Tk()
root.title("DATA ENTRY")

# Set size

wh = 520
ww = 590

# root.resizable(height=False, width=False)

# root.minsize(ww, wh)
# root.maxsize(ww, wh)

# Position in center screen

ws = root.winfo_screenwidth()
hs = root.winfo_screenheight()

# calculate x and y coordinates for the Tk root window
x = (ws/2) - (ww/2)
y = (hs/2) - (wh/2)

root.geometry('%dx%d+%d+%d' % (ww, wh, x, y))

app = Application(root)

root.mainloop()
