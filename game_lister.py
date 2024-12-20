import tkinter
from tkinter import *

from tkinter.ttk import *
from tkinter import messagebox

from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

import os

from time import time, sleep
from datetime import datetime, timedelta

import subprocess as sp

class Application(Frame):

    def __init__(self, master):

        self.master = master
        self.main_container = Frame(self.master)

        # Define the source and target folder variables

        self.origin = os.getcwd()
        self.source = StringVar()
        self.target = ""
        self.initFolders = IntVar()
        self.ftype = IntVar()
        self.pointer = 0
        self.identifier = StringVar()
        self.sheet = StringVar()
        self.sheet_saved = StringVar()
        self.sheetId = StringVar()
        self.sheetId_saved = StringVar()
        self.gameDesc = StringVar()
        self.allMoves = []
        self.allComments = {}
        self.winner = IntVar()
        self.advantage = IntVar()
        self.showFlag = IntVar()
        self.whiteNoteFlag = IntVar()
        self.blackNoteFlag = IntVar()
        self.whiteHasNote = IntVar()
        self.blackHasNote = IntVar()
        self.sheetsList = ['No Selection ']
        self.sheetIdsList = ['No Selection']

        # Create main frame
        self.main_container.grid(column=0, row=0, sticky=(N,S,E,W))

        # Set Label styles
        Style().configure("M.TLabel", font="Courier 20 bold", height="20", foreground="blue", anchor="center")
        Style().configure("B.TLabel", font="Verdana 8", background="white", width="25")
        Style().configure("G.TLabel", font="Verdana 8")
        Style().configure("L.TLabel", font="Courier 40 bold", width="8")
        Style().configure("MS.TLabel", font="Verdana 10" )
        Style().configure("S.TLabel", font="Verdana 8" )
        Style().configure("G.TLabel", font="Verdana 8")

        # Set button styles
        Style().configure("B.TButton", font="Verdana 8", relief="ridge")

        # Set check button styles
        Style().configure("B.TCheckbutton", font="Verdana 8")
        Style().configure("B.TRadiobutton", font="Verdana 8")
        Style().configure("O.TLabelframe.Label", font="Verdana 8", foreground="black")

        # Create widgets
        self.sep_a = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_b = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_c = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_d = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_e = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_f = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_g = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_h = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_i = Separator(self.main_container, orient=HORIZONTAL)
        self.mainLabel = Label(self.main_container, text="GAME MOVES", style="M.TLabel" )

        self.sourceOption = LabelFrame(self.main_container, text=' FILE ', style="O.TLabelframe")
        self.selectSource = Button(self.sourceOption, text="SOURCE", style="B.TButton", command=self.setSource)
        self.sourceLabel = Label(self.sourceOption, text="None", style="B.TLabel" )
        self.sheetOptions = LabelFrame(self.main_container, text=' SHEETS ', style="O.TLabelframe")
        self.selectSheet = OptionMenu(self.sheetOptions, self.sheet, *self.sheetsList, command=self.getSheetIdsList)
        self.report = Button(self.main_container, text="REPORT", style="B.TButton", command=self.reportComments)

        self.idOptions = LabelFrame(self.main_container, text=' IDENTIFIERS ', style="O.TLabelframe")
        self.idList = Listbox(self.idOptions, selectmode='single', width=60, height=5)
        self.iscroller = Scrollbar(self.idOptions, orient=VERTICAL, command=self.idList.yview)
        self.idList.config(font=("Courier New", 8), yscrollcommand=self.iscroller.set)

        self.whiteFrame = LabelFrame(self.main_container, text=' WHITE ', style="O.TLabelframe")
        self.whiteMove = Label(self.whiteFrame, text=" ", style="L.TLabel" )
        self.whiteNote = Checkbutton(self.whiteFrame, text=" Note found", style="B.TCheckbutton", variable=self.whiteNoteFlag)
        self.blackFrame = LabelFrame(self.main_container, text=' BLACK ', style="O.TLabelframe")
        self.blackMove = Label(self.blackFrame, text=" ", style="L.TLabel" )
        self.blackNote = Checkbutton(self.blackFrame, text=" Note found", style="B.TCheckbutton", variable=self.blackNoteFlag)
        
        self.next = Button(self.main_container, text="NEXT", style="B.TButton", command=self.getNextMove)
        self.prev = Button(self.main_container, text="PREV", style="B.TButton", command=self.getPrevMove)
        self.info = Button(self.main_container, text="GAME MOVES AND INFO", style="B.TButton", command=self.displayMovesPanel)
        self.restart = Button(self.main_container, text="RESTART", style="B.TButton", command=self.restartMoves)
        self.start = Button(self.main_container, text="PLAY", style="B.TButton", command=self.startGame)
        self.reset = Button(self.main_container, text="RESET FILE/SHEET OPTIONS", style="B.TButton", command=self.resetProcess)

        self.showAll = Button(self.main_container, text="SHOW ALL", style="B.TButton", command=self.showAllMoves)

        self.exit = Button(self.main_container, text="EXIT", style="B.TButton", command=root.destroy)

        self.progress_bar = Progressbar(self.main_container, orient="horizontal", mode="indeterminate", maximum=50)

        # Position widgets
        self.mainLabel.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.sep_a.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.selectSource.grid(row=0, column=0, padx=5, pady=(5,10), sticky='NSW')
        self.sourceLabel.grid(row=0, column=0, padx=(100,10), pady=(5,10), sticky='NSW')
        self.sourceOption.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='NSW')
        self.selectSheet.grid(row=0, column=0, columnspan=1, padx=5, pady=5, sticky='NSEW')
        self.selectSheet.config(width=30)
        self.sheetOptions.grid(row=2, column=2, columnspan=1, padx=5, pady=5, sticky='NSEW')
        
        self.sep_b.grid(row=3, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.report.grid(row=4, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.sep_c.grid(row=5, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
                
        self.idList.grid(row=0, column=0, columnspan=3, padx=5, pady=5, sticky='W')
        self.iscroller.grid(row=0, column=3, columnspan=1, padx=5, pady=5, sticky='W')
        self.idOptions.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.sep_d.grid(row=7, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.start.grid(row=8, column=0, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.reset.grid(row=8, column=2, columnspan=2, padx=5, pady=5, sticky='NSEW')
        
        self.sep_e.grid(row=9, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.whiteMove.grid(row=0, column=0, padx=5, pady=5, sticky="NSEW")
        self.whiteNote.grid(row=1, column=0, padx=5, pady=5, sticky="NSEW")
        self.whiteFrame.grid(row=10, column=0, columnspan=2, rowspan=1, padx=5, pady=5, sticky="NSEW")
        
        self.blackMove.grid(row=0, column=0, padx=5, pady=5, sticky="NSEW")
        self.blackNote.grid(row=1, column=0, padx=5, pady=5, sticky="NSEW")
        self.blackFrame.grid(row=10, column=2, columnspan=2, rowspan=1, padx=5, pady=5, sticky="NSEW")
                
        self.prev.grid(row=12, column=0, columnspan=2, padx=5, pady=0, sticky='NSEW')
        self.next.grid(row=12, column=2, columnspan=2, padx=5, pady=0, sticky='NSEW')
        self.restart.grid(row=13, column=0, columnspan=2, padx=5, pady=0, sticky='NSEW')
        self.info.grid(row=13, column=2, columnspan=2, padx=5, pady=0, sticky='NSEW')
        
        self.sep_f.grid(row=14, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.exit.grid(row=15, column=0, columnspan=4, padx=5, pady=0, sticky='NSEW')

        self.sep_g.grid(row=16, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.progress_bar.grid(row=17, column=0, columnspan=4, padx=5, pady=0, sticky='NSEW')

        self.processControl(0)

    def setSource(self):

        pathname = askopenfilename()

        self.source.set(pathname)

        try:
            if self.source.get().endswith(".xlsx") or self.source.get().endswith(".xls"):
                self.source.set(pathname)
                self.sourceLabel["text"] = self.source.get().split("/")[-1]
                self.getSheetList()

            else:
                messagebox.showerror("Invalid file selected", "Invalid file type was selected. Please select again.")
                self.source.set('')
        except:
            pass

    def getSheetList(self):

        wb = load_workbook(self.source.get())

        # self.sheetsList = []
        self.sheetsList = ['No Selection ']

        for ws in wb.worksheets:
            self.sheetsList.append(ws.title)

        menu = self.selectSheet['menu']
        menu.delete(0, 'end')

        for sh in self.sheetsList:
            menu.add_command(label=sh, command=lambda value=sh: self.getSheetIdsList(value))


    def getSheetIdsList(self, sheet):

        self.sheet.set(sheet)
        self.sheet_saved.set(sheet)
        
        self.getGameIds(sheet)

        self.idList.delete(0, END)
        for id in self.sheetIdsList:
            self.idList.insert(END, id)

    def getOpeningIds(self, sheet):

        wb = load_workbook(self.source.get())
        ws = wb[sheet]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']

        self.sheetIdsList = []

        count = 1
        eol = False
        
        while True:
            
            for col in cola:
                
                cell = col + str(count)
                
                if ws[cell].value == 'END':
                    eol = True
                    break
                    
                if ws[cell].value:

                    bgColor = ws[cell].fill.fgColor.index
                    adv = self.checkAdvantage(bgColor)

                    try:
                        id = ws[cell].value + ' (' + adv + ')' + ' - ' + ws[cell].comment.text.rstrip()
                    except:
                        id = ws[cell].value + ' (' + adv + ')'

                    self.sheetIdsList.append(id)
                
            if eol:
                break
                
            count += 21

    def getGameIds(self, sheet):
        
        wb = load_workbook(self.source.get())
        
        ws = wb[sheet]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']

        self.sheetIdsList = []

        count = 1
        
        for col in cola:
            
            cell = col + str(count)

            if ws[cell].value:
                
                bgColor = ws[cell].fill.fgColor.index
                adv = self.checkAdvantage(bgColor)

                try:
                    id = ws[cell].value + ' (' + adv + ')' + ' - ' + ws[cell].comment.text.rstrip()
                except:
                    id = ws[cell].value + ' (' + adv + ')'
                    
                self.sheetIdsList.append(id)

    def startGame(self):

        if self.sheet.get():
            pass
        else: 
            messagebox.showerror("Select sheet", "Please select sheet from workbook.")
            return 

        if self.idList.curselection():
            self.sheetId.set(self.idList.get(self.idList.curselection()[0])[:6])
        else: 
            messagebox.showerror("Select sheet ID", "Please select sheet ID from worksheet.")
            return 

        self.loadGameMoves()
        self.postFirstMove()
        self.processControl(1)

    def locateMoves(self, length):

        wb = load_workbook(self.source.get())
        ws = wb[self.sheet.get()]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
        colb = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

        found_tag = False

        for ca, cb in zip(cola, colb):
            cell = ca + '1'
            
            if ws[cell].value == self.sheetId.get()[:length]:    
                self.workSheet = ws 
                self.whiteColumn = ca 
                self.blackColumn = cb 
                return True 
        
        return False 
            
    def loadGameMoves(self):

        if self.locateMoves(6):
            pass 
        else: 
            return 

        self.allMoves = []
        count = 2
        
        while True:
            
            cell = self.whiteColumn + str(count)
            
            if self.workSheet[cell].value:
                self.allMoves.append(self.workSheet[cell].value) 
            else:
                break

            cell = self.blackColumn + str(count)
            
            if self.workSheet[cell].value:
                self.allMoves.append(self.workSheet[cell].value)
            else:
                break
            
            count += 1

        self.pointer = 0

        d_cell = self.blackColumn + '1'
        
        try:
            self.gameDesc.set(self.workSheet[d_cell].comment.text.rstrip())
        except:
            self.gameDesc.set('No description')

    def postFirstMove(self):

        move = self.allMoves[self.pointer]
        self.whiteMove["text"] = move
        self.blackMove["text"] = ""

        self.checkForComments()

    def getNextMove(self):

        self.checkCommentNotes()

        if self.pointer + 1== len(self.allMoves):
            messagebox.showinfo("Last moves", "Last moves already displayed.")
            return

        self.pointer += 1

        move = self.allMoves[self.pointer]

        if self.pointer % 2 == 1:
            self.blackMove["text"] = move
        else:
            self.whiteMove["text"] = move
            self.blackMove["text"] = ""

        self.checkForComments()

    def getPrevMove(self):

        if self.pointer == 0:
            messagebox.showinfo("First moves", "First moves already displayed.")
            return

        self.pointer -= 1

        move = self.allMoves[self.pointer]

        if self.pointer % 2 == 1:
            self.blackMove["text"] = move 

            white = self.allMoves[self.pointer - 1]
            self.whiteMove["text"] = white
        else:
            self.whiteMove["text"] = move
            self.blackMove["text"] = ""

        self.checkForComments()

    def get_tag(self, tag):
        
        wb2 = load_workbook(self.source.get())
        
        try:
            ws_index = wb2['Index']
            
            count = 2
            
            while True:
                
                cell = 'A' + str(count)
                
                if ws_index[cell].value == tag:
                    opening = ws_index['B' + str(count)].value
                    white = ws_index['D' + str(count)].value
                    black = ws_index['E' + str(count)].value

                    return opening, white, black
                    
                count += 1          
        except:
            return self.sheet.get(), "Player 1", "Player 2"
        
    def displayMovesPanel(self):

        Style().configure("PS.TLabel", font="Verdana 8", height="50" )
        self.popMoves = Toplevel(self.main_container)
        self.popMoves.title(self.sheetId.get())

        self.pop_a = Separator(self.popMoves, orient=HORIZONTAL)
        self.pop_b = Separator(self.popMoves, orient=HORIZONTAL)
        self.pop_c = Separator(self.popMoves, orient=HORIZONTAL)
        self.pop_d = Separator(self.popMoves, orient=HORIZONTAL)
        self.pop_e = Separator(self.popMoves, orient=HORIZONTAL)

        self.popPlayers = Label(self.popMoves, text=" ", style="S.TLabel" )
        self.popOpening = Label(self.popMoves, text=" ", style="S.TLabel" )
        self.popDescription = Text(self.popMoves, width="42", height="5" )

        self.upddesc = Button(self.popMoves, text="UPDATE", style="B.TButton", command=self.updateDescription)

        self.moveListFrame = LabelFrame(self.popMoves, text=' MOVES LIST ', style="O.TLabelframe")
        self.moveList = Listbox(self.moveListFrame, width=38, height=8)
        self.scroller = Scrollbar(self.moveListFrame, orient=VERTICAL, command=self.moveList.yview)
        self.moveList.config(font=("Courier New", 10), yscrollcommand=self.scroller.set)
        
        self.popOpening.configure(font=("Courier New", 10))
        self.popPlayers.configure(font=("Courier New", 10))

        self.close = Button(self.popMoves, text="CLOSE", style="B.TButton", command=self.hidePopup)

        self.popPlayers.grid(row=1, column=0, columnspan=3, padx=5, pady=1, sticky="NSEW")
        self.popOpening.grid(row=2, column=0, columnspan=2, padx=5, pady=1, sticky="NSEW")
        
        self.pop_a.grid(row=3, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")        

        self.popDescription.grid(row=4, column=0, columnspan=2, padx=5, pady=1, sticky="NSEW")
        self.upddesc.grid(row=5, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

        self.pop_b.grid(row=6, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")        

        self.moveList.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky='NSEW')
        self.scroller.grid(row=0, column=2, columnspan=1, padx=5, pady=5, sticky='NSEW')
        self.moveListFrame.grid(row=7, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

        self.pop_c.grid(row=8, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

        self.close.grid(row=9, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

        opening, white, black = self.get_tag(self.sheetId.get())

        self.popPlayers["text"] = white + " vs. " + black
        self.popDescription.insert(INSERT, self.gameDesc.get())

        self.loadGameList()
        self.showFlag.set(0)

        if self.advantage.get() == 2:
            self.popOpening["text"] = opening + " (W)"
        elif self.advantage.get() == 1: 
            self.popOpening["text"] = opening + " (B)"
        else:
            self.popOpening["text"] = opening + " (D)"

        ph = 420
        pw = 360

        self.popMoves.maxsize(pw, ph)
        self.popMoves.minsize(pw, ph)

        ws = self.popMoves.winfo_screenwidth()
        hs = self.popMoves.winfo_screenheight()

        x = (ws/2) - (pw/2) + 500
        y = (hs/2) - (ph/2)

        self.popMoves.geometry('%dx%d+%d+%d' % (pw, ph, x, y))

        self.info["state"] = DISABLED
        
    def loadGameList(self):

        self.moveList.delete(0, END)

        cell = self.whiteColumn + '1'
        
        bgColor = self.workSheet[cell].fill.fgColor.index
        self.checkAdvantage(bgColor)

        self.locateMoves(6)

        count = 2
        
        while True:
            
            cell = self.whiteColumn + str(count)
            
            if self.workSheet[cell].value:
                white_move = self.workSheet[cell].value
            else:
                break

            cell = self.blackColumn + str(count)
            
            if self.workSheet[cell].value:
                black_move = self.workSheet[cell].value
            else:
                self.moveList.insert(END, '{:2d}'.format(count - 1) + '. ' + white_move)
                break
            
            self.moveList.insert(END, '{:2d}'.format(count - 1) + '. ' + white_move.ljust(6) + '  -  ' + black_move)

            count += 1

    def checkAdvantage(self, fill):

        self.advantage.set(1)
        if fill == "FF00FF00":          # green
            self.advantage.set(2)
            return 'W'
        elif fill == "FFFFFF00":        # yellow
            self.advantage.set(1)
            return 'B'
        else:                           # turquise
            self.advantage.set(0)
            return 'D'

    def checkForComments(self):

        row = int(self.pointer / 2) + self.pointer % 2

        self.whiteNoteFlag.set(0)
        self.blackNoteFlag.set(0)

        ''' check for comments first
        '''
        if self.pointer % 2 == 0: # white move
            w_cell = self.whiteColumn + str(row + 2) 

            try:
               temp = self.workSheet[w_cell].comment.text.rstrip()
               self.whiteNoteFlag.set(1)
               self.whiteHasNote.set(1)

            except Exception as e:

                fgColor = self.workSheet[w_cell].fill.fgColor.index

                if fgColor == "00000000":
                    self.whiteNoteFlag.set(0)
                    self.whiteHasNote.set(0)
                else:    
                    self.whiteNoteFlag.set(1)
                    self.whiteHasNote.set(1)
            
        else:
            b_cell = self.blackColumn + str(row + 1)

            try:
                temp = self.workSheet[b_cell].comment.text.rstrip()
                self.blackNoteFlag.set(1)
                self.blackHasNote.set(1)

            except Exception as e:

                fgColor = self.workSheet[b_cell].fill.fgColor.index

                if fgColor == "00000000":
                    self.blackNoteFlag.set(0)
                    self.blackHasNote.set(0)
                else:    
                    self.blackNoteFlag.set(1)
                    self.blackHasNote.set(1)


    def checkCommentNotes(self):

        ''' check and record comments only after black move
        '''

        wb = load_workbook(self.source.get())
        ws = wb[self.sheet.get()]
 
        if self.pointer % 2 == 1: # if black just moved

            row = int(self.pointer / 2) 

            cell = self.whiteColumn +  str(row + 2)

            update_wb = False 
            if self.whiteNoteFlag.get() and not self.whiteHasNote.get():
                ws[cell].fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
                update_wb = True

            cell = self.blackColumn +  str(row + 1)

            if self.blackNoteFlag.get() and not self.blackHasNote.get():
                ws[cell].fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
                update_wb = True

            if update_wb:
                wb.save(self.source.get())

    def showAllMoves(self):

        if self.showFlag.get() == 0:
            self.showFlag.set(1)
            self.showAll.configure(text="HIDE ALL")
            ph = 380
        else:
            self.showFlag.set(0)
            self.showAll.configure(text="SHOW ALL")
            ph = 210

        pw = 410

        self.popMoves.maxsize(pw, ph)
        self.popMoves.minsize(pw, ph)

        ws = self.popMoves.winfo_screenwidth()
        hs = self.popMoves.winfo_screenheight()

        x = (ws/2) - (pw/2)
        y = (hs/2) - (ph/2)

        self.popMoves.geometry('%dx%d+%d+%d' % (pw, ph, x, y))

        self.processControl(0)

    def updateDescription(self):

        wb = load_workbook(self.source.get())
        ws = wb[self.sheet.get()]

        cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
        colb = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

        for ca, cb in zip(cola, colb):
            
            cell = ca + '1'
            
            if ws[cell].value == self.sheetId.get()[:6]:    

                d_cell = cb + '1'

                comment = Comment(self.popDescription.get(1.0, END), "")

                ws[d_cell].comment = comment

                wb.save(self.source.get())

                self.gameDesc.set(ws[d_cell].comment.text.rstrip())


    def hidePopup(self):

        self.info["state"] = NORMAL
        self.popMoves.destroy()

    def restartMoves(self):

        res = messagebox.askquestion(title="Restart moves?", message="Do you want to restart game/opening?")

        if res == 'no':
            return

        self.pointer = 0
        self.postFirstMove()

        messagebox.showinfo("Game restarted", "Game has restarted, first move shown.") 

    def processControl(self, mode):
        ''' enable/disable buttons as needed
        '''

        if mode:
            
            self.prev["state"] = NORMAL
            self.next["state"] = NORMAL
            self.info["state"] = NORMAL
            self.restart["state"] = NORMAL

        else:

            self.prev["state"] = DISABLED
            self.next["state"] = DISABLED
            self.info["state"] = DISABLED
            self.restart["state"] = DISABLED


    def reportComments(self):

        if self.source.get() and self.sheet.get():

            report_file = open('report.txt', 'w')

            wb = load_workbook(self.source.get())
            ws = wb[self.sheet.get()]

            cola = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
            colb = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

            for ca, cb in zip(cola, colb):
                    
                cell = ca + '1'

                if ws[cell].value == 'END':
                    break

                open_id = ws[cell].value

                if open_id:

                    cell = cb + '1'

                    try:
                        report_line = open_id + ' ' + ' - '.join(ws[cell].comment.text.rstrip().split('\n'))
                    except:
                        report_line = open_id + ' No comment'

                    report_file.write(report_line+'\n')

            report_file.close()

            res = messagebox.askquestion(title="View report?", message="Do you want to view report after creation?")

            if res == 'yes':

                work_dir = os.getcwd()

                # temp_dir = os.path.dirname('report.txt')

                # os.chdir(temp_dir)

                os.system(str(os.path.basename('report.txt')))

                os.chdir(work_dir)


        else:

            messagebox.showerror("No source and sheet","No valid source or sheet selected for report")


    def resetProcess(self):
        ''' reset labels, lists and flags
        '''
        
        res = messagebox.askquestion(title="Reset process?", message="Do you want to reset selections?")

        if res == 'no':
            return

        os.chdir(self.origin)
        self.sourceLabel["text"] = "None"
        self.source.set("")

        self.idList.delete(0, END)

        menu = self.selectSheet['menu']
        menu.delete(0, 'end')

        self.sheetsList = ['No Selection ']
        for sh in self.sheetsList:
            menu.add_command(label=sh, command=lambda value=sh: self.getSheetIdsList(value))

        self.processControl(0)

        self.sheet.set('No Selection')
        self.sheetId.set('')

        self.whiteMove["text"] = ""
        self.blackMove["text"] = ""


root = Tk()
root.title("GAMES MOVES")

# Set size

wh = 610
ww = 590

root.resizable(height=False, width=False)

root.minsize(ww, wh)
root.maxsize(ww, wh)

# Position in center screen

ws = root.winfo_screenwidth()
hs = root.winfo_screenheight()

# calculate x and y coordinates for the Tk root window
x = (ws/2) - (ww/2)
y = (hs/2) - (wh/2)

root.geometry('%dx%d+%d+%d' % (ww, wh, x, y))

app = Application(root)

root.mainloop()
